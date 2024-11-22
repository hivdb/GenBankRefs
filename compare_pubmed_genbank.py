import pandas as pd
from collections import defaultdict
import Levenshtein
from collections import Counter
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def search_access_prefix(pubmed, accession_prefix_list):
    found = pd.DataFrame()

    for acc_prefix in accession_prefix_list:
        result = pubmed[pubmed['GenBank'].apply(
            lambda x: acc_prefix.lower() in str(x).lower())]
        if not result.empty:
            found = pd.concat([found, result])

    return found


def main():
    genbank_file = 'CCHF_Combined_11_21.xlsx'
    pubmed_file = 'ReferenceSummary_Nov22.xlsx'
    genbank_ref_file = 'ReferenceSummary_Genbank_Nov20.xlsx'

    genbank = pd.read_excel(genbank_file)
    pubmed = pd.read_excel(pubmed_file)
    pubmed = pd.concat([pubmed, pd.read_excel(genbank_ref_file)])

    pubmed = pubmed[((pubmed['Reviewer(s) Seq'] == 'Yes') & (
        pubmed['GPT seq (Y/N)'] == 'Yes')) | (pubmed['Resolve'] == 'Yes')]

    # print(len(pubmed))

    match_by_title = []
    match_by_pmid = []
    match_by_acc = []
    matched_pubmed_indices = []
    genbank_unmatch = []

    for index, row in genbank.iterrows():
        pmid = row['pmid']

        result = pubmed[pubmed['PMID'] == pmid]

        if not result.empty:
            matched_pubmed_indices.extend(result.index.tolist())
            match_by_pmid.append([row, result])
            continue

        title = row['title']

        result = pubmed[pubmed['Title'].apply(
            lambda x: Levenshtein.distance(x.lower(), title.lower()) < 5)]

        if not result.empty:
            matched_pubmed_indices.extend(result.index.tolist())
            match_by_title.append([row, result])
            continue

        accession_list = row['accession']
        accession_prefix_list = set([
            a.strip()[:6]
            for a in accession_list.split(',')
        ])

        # if 'JX908640' in accession_list:
        #     print(accession_prefix_list)

        result = search_access_prefix(pubmed, accession_prefix_list)
        if not result.empty:
            matched_pubmed_indices.extend(result.index.tolist())
            match_by_acc.append([row, result])
            continue

        genbank_unmatch.append(row)

    pubmed.to_excel('Pubmed.xlsx')

    genbank_match = match_by_title + match_by_pmid + match_by_acc

    print("Genbank match by pmid", len(match_by_pmid))
    print("Genbank match by title", len(match_by_title))
    print("Genbank match by acc", len(match_by_acc))
    print('Genbank Matched:', len(genbank_match))

    pubmed_match = match_pubmed2genbank(genbank_match)

    print("Pubmed match by pmid", len(match_pubmed2genbank(match_by_pmid)))
    print("Pubmed match by title", len(match_pubmed2genbank(match_by_title)))
    print("Pubmed match by acc", len(match_pubmed2genbank(match_by_acc)))
    print('Pubmed Matched:', len(pubmed_match))

    pubmed_unmatch = pubmed.drop(index=matched_pubmed_indices)

    # pubmed_unmatch.to_excel('pubmed_unmatch.xlsx')
    # pd.DataFrame(genbank_unmatch).to_excel('genbank_unmatch.xlsx')

    combined, columns = combine(pubmed_match, pubmed_unmatch, genbank_unmatch)
    combined.to_excel('combined.xlsx', index=False, columns=columns)

    format_table('combined.xlsx')

    summarize_combined_data(combined)

    # Get PubMed data summary
    summarize_pubmed_data(pubmed)
    # Get Genbank data summary
    summarize_genbank_data(genbank)


def categorize_host(entry):
    if pd.isnull(entry):  # Handle NaN or null values
        return 'NA'
    # Remove all spaces and lowercase
    entry = entry.replace(" ", "").strip().lower()
    if 'tick' in entry or 'Tick' in entry:  # Check if "tick" is present in the processed string
        return 'Ticks'
    if 'serum' in entry or 'blood' in entry or "plasma" in entry:
        return 'blood'
    return entry


def middle_year(entry):
    entry = str(entry)  # Ensure the entry is a string
    if entry == "None":
        return "NA"
    # Remove whitespace and replace different dash types with a standard "-"
    entry = re.sub(r"[–—]", "-", entry.replace(" ", ""))

    if '-' in entry and ',' not in entry:  # Handle ranges like "2012-2013"
        try:
            start, end = map(int, entry.split('-'))
            return (start + end) // 2
        except ValueError:
            # print(entry)
            return "NA"  # Handle malformed ranges gracefully
    elif ',' in entry:  # Handle lists like "1985,1990,2000"
        try:
            years = [int(year) for year in entry.split(',')]
            return sum(years) // len(years)
        except ValueError:
            # print(entry)
            return "NA"  # Handle malformed lists gracefully
    else:  # Handle single years
        try:
            return int(entry)
        except ValueError:
            # print(entry)
            return "NA"


def clean_entry(entry):
    if pd.isnull(entry):  # Handle NaN values
        return 'NA'
    return str(entry).strip()


def summarize_pubmed_data(df):
    print("PubMed")
    publish_year = count_number([v for i, v in df.iterrows()], 'Year')
    print('Publish Year')
    print(publish_year)
    print('=' * 40)

    journals = count_number([v for i, v in df.iterrows()], 'Journal')
    print('Journals')
    print(journals)
    print('=' * 40)

    methods = count_number([v for i, v in df.iterrows()], 'SeqMethod')
    print('Seq method')
    print(methods)
    print('=' * 40)

    ffiltered_num_seq_rows = [
        {'NumSeqs': int(v['NumSeqs'])}
        for i, v in df.iterrows() if not pd.isnull(v['NumSeqs'])]
    num_seq = count_number(ffiltered_num_seq_rows, 'NumSeqs')
    print('NumSeq', num_seq)
    print('=' * 40)

    df['CleanedHosts'] = df['Host'].apply(categorize_host)
    hosts = count_number([v for i, v in df.iterrows()], 'CleanedHosts')
    print('Host')
    print(hosts)
    print('=' * 40)

    df['CleanedIsolate'] = df['IsolateType'].apply(categorize_host)
    specimen = count_number([v for i, v in df.iterrows()], 'CleanedIsolate')
    print('Specimen')
    print(specimen)
    print('=' * 40)

    df['MiddleYear'] = df['SampleYr'].apply(middle_year)
    year = count_number([v for i, v in df.iterrows()], 'MiddleYear')
    print('MiddleYear')
    print(year)
    print('=' * 40)

    cleaned_country_values = [
        {'Country': clean_entry(v['Country'])} for i, v in df.iterrows() if 'Country' in v
    ]
    country = count_number(cleaned_country_values, 'Country')
    print('Country')
    print(country)
    print('=' * 40)

    gene = count_number([v for i, v in df.iterrows()], 'Gene')
    print('Gene')
    print(gene)
    print('=' * 40)


def get_most_recent(years):
    if pd.isnull(years) or years == 'nan':  # Check for NaN or 'nan'
        return 'nan'
    # Split the string into individual years, clean them, and convert to integers
    year_list = [int(y.strip()) for y in str(years).split(',')]
    # Return the most recent year
    return max(year_list)


def remove_parenthesis(entry):
    # Remove everything in parentheses
    entry = re.sub(r"\s*\([^)]*\)", "", entry)
    # Extract the main part of the entry (excluding numbers)
    match = re.match(r"^[^0-9]+", entry)  # Correctly define 'match'
    # Return the matched portion if it exists; otherwise, return the cleaned entry
    return match.group().strip() if match else entry


def calculate_average_length(entry):
    # Split by commas and process each component
    components = entry.split(',')
    total_length = 0
    total_frequency = 0

    for component in components:
        # Remove leading/trailing whitespace
        component = component.strip()

        # Check for range
        range_match = re.match(r"(\d+)-(\d+)\s*\((\d+)\)", component)
        if range_match:
            start, end, freq = map(int, range_match.groups())
            avg_length = (start + end) / 2  # Average the range
            frequency = freq
        else:
            # Match single value with frequency
            value_match = re.match(r"(\d+)\s*\((\d+)\)", component)
            if value_match:
                value, freq = map(int, value_match.groups())
                avg_length = value
                frequency = freq
            else:
                # Skip invalid entries
                continue

        # Update totals
        total_length += avg_length * frequency
        total_frequency += frequency

    # Calculate weighted average

    return round(total_length / total_frequency) if total_frequency > 0 else 0


def most_frequent_range(entry):
    # Split by commas to process each range and frequency
    components = entry.split(',')
    max_frequency = 0
    most_frequent = None

    for component in components:
        # Match ranges and frequencies
        match = re.match(r"([\d%\-]+)\s*\((\d+)\)", component.strip())
        if match:
            range_, freq = match.groups()
            freq = int(freq)
            # Update the most frequent range
            if freq > max_frequency:
                max_frequency = freq
                most_frequent = range_

    return most_frequent


def summarize_genbank_data(df):
    print("GenBank")
    df['MostRecentYear'] = df['year'].apply(get_most_recent)
    publish_year = count_number(
        [v for i, v in df.iterrows()], 'MostRecentYear')
    print('Publish Year')
    print(publish_year)
    print('=' * 40)

    journal_values = [row['journal'].split(',')[0].strip()
                      for _, row in df.iterrows() if 'journal' in row and pd.notnull(row['journal'])]
    cleaned_entries = [remove_parenthesis(entry) for entry in journal_values]
    journals = count_number([{'journal': value}
                            for value in cleaned_entries], 'journal')
    print('Journals')
    print(journals)
    print('=' * 40)

    num_seq = [len(v) for v in df['accession']]
    counter = dict(Counter(num_seq))
    num_seqs = ','.join([
        f'{k} ({v})'
        for k, v in sorted(counter.items(), key=lambda x: int(x[-1]), reverse=True)
    ])
    print('NumSeq', num_seqs)
    print('=' * 40)

    # Preprocess the 'Hosts' field to clean and normalize the data
    host_values = [row['Hosts'].strip()
                   for _, row in df.iterrows() if 'Hosts' in row and pd.notnull(row['Hosts'])]
    cleaned_entries = [remove_parenthesis(entry) for entry in host_values]
    host = count_number([{'Hosts': value}
                        for value in cleaned_entries], 'Hosts')
    print('Host')
    print(host)
    print('=' * 40)

    specimen_values = [row['Specimens'].strip()
                       for _, row in df.iterrows() if 'Hosts' in row and pd.notnull(row['Specimens'])]
    cleaned_entries = [remove_parenthesis(entry) for entry in specimen_values]
    host = count_number([{'Specimens': value}
                        for value in cleaned_entries], 'Specimens')
    print('Specimens')
    print(host)
    print('=' * 40)

    specimen_values = [row['RecordYears'].strip()
                       for _, row in df.iterrows() if 'Hosts' in row and pd.notnull(row['RecordYears'])]
    # print(specimen_values)
    cleaned_entries = [remove_parenthesis(entry) for entry in specimen_values]
    host = count_number([{'RecordYears': value}
                        for value in cleaned_entries], 'RecordYears')
    print('RecordYears')
    print(host)
    print('=' * 40)

    country_values = [row['Countries'].strip()
                      for _, row in df.iterrows() if 'Hosts' in row and pd.notnull(row['Countries'])]
    # print(specimen_values)
    cleaned_entries = [remove_parenthesis(entry) for entry in country_values]
    host = count_number([{'Countries': value}
                        for value in cleaned_entries], 'Countries')
    print('Countries')
    print(host)
    print('=' * 40)

    gene_values = [row['Gene'].strip()
                   for _, row in df.iterrows() if 'Hosts' in row and pd.notnull(row['Gene'])]
    # print(specimen_values)
    cleaned_entries = [remove_parenthesis(entry) for entry in gene_values]
    host = count_number([{'Gene': value} for value in cleaned_entries], 'Gene')
    print('Gene')
    print(host)
    print('=' * 40)

    df['AverageAlignLength'] = df['AlignLens'].apply(calculate_average_length)
    host = count_number([v for i, v in df.iterrows()], 'AverageAlignLength')
    print('AlignLens')
    print(host)
    print('=' * 40)

    df['MostFrequentRange'] = df['PcntIDs'].apply(most_frequent_range)
    host = count_number([v for i, v in df.iterrows()], 'MostFrequentRange')
    print('MostFrequentRange')
    print(host)
    print('=' * 40)


def summarize_combined_data(combined):
    matches = combined[(combined['match'] == 'Yes')]
    countries = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'Countries (GB)')
    print('Countries')
    print(countries)
    print('=' * 40)

    year = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'SampleYr (GB)')
    print("Sample year")
    print(year)
    print('=' * 40)

    host = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'Hosts (GB)')
    print('Host')
    print(host)
    print('=' * 40)

    specimens = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'Specimen (GB)')
    print('Specimens')
    print(specimens)
    print('=' * 40)

    genes = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'Genes (GB)')
    print('Genes')
    print(genes)
    print('=' * 40)

    aligns = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'AlignLens (GB)')
    print('AlignLens')
    print(aligns)
    print('=' * 40)

    pcnt_ident = merge_genbank_list_columns(
        [v for i, v in matches.iterrows()], 'PcntIDs (GB)')
    print('PcntIDs')
    print(pcnt_ident)
    print('=' * 40)

    publish_year = count_number([v for i, v in matches.iterrows()], 'Year')
    print('Publish Year')
    print(publish_year)
    print('=' * 40)

    journals = count_number([v for i, v in matches.iterrows()], 'Journal')
    print('Journals')
    print(journals)
    print('=' * 40)

    methods = count_number(
        [v for i, v in matches.iterrows()], 'SeqMethod (PM)')
    print('Seq method')
    print(methods)
    print('=' * 40)

    num_seq = sum([int(v['NumSeqs (GB)']) for i, v in matches.iterrows()])
    print('NumSeq', num_seq)
    print('=' * 40)


def match_pubmed2genbank(genbank_match):
    pubmed_match = defaultdict(dict)
    for g, pubmed_list in genbank_match:
        for r, i in pubmed_list.iterrows():
            pubmed_match[r]['pubmed'] = i
            if 'genbank' not in pubmed_match[r]:
                pubmed_match[r]['genbank'] = []

            pubmed_match[r]['genbank'].append(g)

    pubmed_match = [
        [v['pubmed'], v['genbank']]
        for v in pubmed_match.values()
    ]

    return pubmed_match


def split_value_count(value_count):

    value_count = value_count[::-1]

    count = value_count[value_count.find(
        ')') + 1: value_count.find('(')][::-1].strip()
    value = value_count[value_count.find('(') + 1:][::-1].strip()

    return value, count


def merge_genbank_list_columns(genbank_list, key):

    column_values = [genbank[key] for genbank in genbank_list]
    value_count_list = [
        j.strip()
        for i in column_values
        for j in i.split(',')
    ]
    value_count_list = [
        split_value_count(j)
        for j in value_count_list
    ]

    result = defaultdict(int)
    for value, count in value_count_list:
        result[
            value
            if value != 'NA' else value
        ] += int(count)

    return ','.join([
        f'{k} ({v})'
        for k, v in sorted(result.items(), key=lambda x: int(x[-1]), reverse=True)
    ])


def count_number(rows, key):
    column_values = [genbank[key] for genbank in rows]
    counter = dict(Counter(column_values))

    return ','.join([
        f'{k} ({v})'
        for k, v in sorted(counter.items(), key=lambda x: int(x[-1]), reverse=True)
    ])


def combine(pubmed_match, pubmed_unmatch, genbank_unmatch):

    result = []
    for pubmed, genbank_list in pubmed_match:
        row = {
            'Authors': pubmed['Authors'],
            'Title': pubmed['Title'],
            'Journal': pubmed['Journal'],
            'Year': pubmed['Year'],
            'PMID': pubmed['PMID'],

            'Reviewer(s) Seq': pubmed['Reviewer(s) Seq'],
            'GPT seq (Y/N)': pubmed['GPT seq (Y/N)'],
            'Resolve': pubmed['Resolve'],
            'match': 'Yes',

            'Viruses (PM)': pubmed['Viruses'],
            'NumSeqs (PM)': pubmed['NumSeqs'],
            'Hosts (PM)': pubmed['Host'],
            'Specimen (PM)': pubmed['IsolateType'],

            'SampleYr (PM)': pubmed['SampleYr'],
            'Countries (PM)': pubmed['Country'],
            'Genes (PM)': pubmed['Gene'],
            'SeqMethod (PM)': pubmed['SeqMethod'],
            'CloneMethod (PM)': pubmed['CloneMethod'],
            'GenBank (PM)': pubmed['GenBank'],

            'Viruses (GB)': merge_genbank_list_columns(genbank_list, 'Organisms'),
            'NumSeqs (GB)': len([
                i
                for genbank in genbank_list
                for i in genbank['accession'].split(',')
            ]),
            'Hosts (GB)': merge_genbank_list_columns(genbank_list, 'Hosts'),

            'Specimen (GB)': merge_genbank_list_columns(genbank_list, 'Specimens'),
            'SampleYr (GB)':  merge_genbank_list_columns(genbank_list, 'IsolateYears'),
            'Countries (GB)': merge_genbank_list_columns(genbank_list, 'Countries'),
            'Genes (GB)': merge_genbank_list_columns(genbank_list, 'Gene'),
            'SeqMethod (GB)': '',
            'CloneMethod (GB)': '',
            'GenBank (GB)': ', '.join([i.strip().split('.')[0]
                                       for genbank in genbank_list
                                       for i in genbank['accession'].split(',')
                                       ]),
            'AlignLens (GB)': merge_genbank_list_columns(genbank_list, 'AlignLens'),
            'PcntIDs (GB)': merge_genbank_list_columns(genbank_list, 'PcntIDs'),
        }

        result.append(row)

    for r, pubmed in pubmed_unmatch.iterrows():
        row = {
            'Authors': pubmed['Authors'],
            'Title': pubmed['Title'],
            'Journal': pubmed['Journal'],
            'Year': pubmed['Year'],
            'PMID': pubmed['PMID'],

            'Reviewer(s) Seq': pubmed['Reviewer(s) Seq'],
            'GPT seq (Y/N)': pubmed['GPT seq (Y/N)'],
            'Resolve': pubmed['Resolve'],

            'Viruses (PM)': pubmed['Viruses'],
            'NumSeqs (PM)': pubmed['NumSeqs'],
            'Hosts (PM)': pubmed['Host'],
            'Specimen (PM)': pubmed['IsolateType'],
            'SampleYr (PM)': pubmed['SampleYr'],
            'Countries (PM)': pubmed['Country'],
            'Genes (PM)': pubmed['Gene'],
            'SeqMethod (PM)': pubmed['SeqMethod'],
            'CloneMethod (PM)': pubmed['CloneMethod'],
            'GenBank (PM)': pubmed['GenBank'],
        }

        result.append(row)

    for genbank in genbank_unmatch:
        numSeqs = len(genbank['accession'].split(','))
        row = {
            'Authors': genbank['authors'],
            'Title': genbank['title'],
            'Journal': genbank['journal'],
            'Year': genbank['year'],
            'PMID': genbank['pmid'],
            'Viruses (GB)': genbank['Organisms'],
            'NumSeqs (GB)': numSeqs,
            'Hosts (GB)': genbank['Hosts'],
            'Specimen (GB)': genbank['Specimens'],
            'SampleYr (GB)': genbank['IsolateYears'],
            'Countries (GB)': genbank['Countries'],
            'Genes (GB)': genbank['Gene'],
            'SeqMethod (GB)': '',
            'CloneMethod (GB)': '',
            'GenBank (GB)': ', '.join([
                i.strip().split('.')[0]
                for i in genbank['accession'].split(',')
            ]),
            'AlignLens (GB)': genbank['AlignLens'],
            'PcntIDs (GB)': genbank['PcntIDs'],
        }

        result.append(row)

    columns = [
        'Authors',
        'Title',
        'Journal',
        'Year',
        'PMID',

        'Reviewer(s) Seq',
        'GPT seq (Y/N)',
        'Resolve',
        'match',

        'Viruses (PM)',
        'Viruses (GB)',

        'NumSeqs (PM)',
        'NumSeqs (GB)',

        'Hosts (PM)',
        'Hosts (GB)',

        'Specimen (PM)',
        'Specimen (GB)',

        'SampleYr (PM)',
        'SampleYr (GB)',

        'Countries (PM)',
        'Countries (GB)',

        'Genes (PM)',
        'Genes (GB)',

        'SeqMethod (PM)',
        'SeqMethod (GB)',

        'CloneMethod (PM)',
        'CloneMethod (GB)',

        'GenBank (PM)',
        'GenBank (GB)',

        'AlignLens (GB)',
        'PcntIDs (GB)',
    ]

    for i in result:
        for c in columns:
            if c not in i:
                i[c] = ''

    return pd.DataFrame(result), columns


def format_table(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 20

    wb.save(excel_file)


if __name__ == '__main__':
    main()
