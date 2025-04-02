import pandas as pd
from collections import defaultdict
import Levenshtein
import pandas as pd
import re
from operator import itemgetter
from collections import Counter

from Utilities import count_number
from Utilities import int_sorter
from Utilities import get_values_of_value_count_list
from Utilities import create_binnned_year

from summarize_genbank import summarize_genbank_by_ref
from summarize_pubmed import summarize_pubmed_data
from summarize_genbank import summarize_genbank_full_genome
from summarize_genbank import summarize_genbank_by_seq

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from DataFrameLogic import merge_feature_rows


def match_pubmed_GB(
        pubmed, genbank_ref, genbank_feature, genbank_genes, virus_obj):
    """
    matches PubMed literature with GenBank references, features, and genes,
    then processes and logs the results. Chord Diagram currently commented out

    Output:
    pubmed: Dataframe with added columns
    pubmed_match: Subset of pubmed Dataframe that have a matching GenBank record.

    Excel Files:
    - pubmed_unmatch_file: Contains PubMed records without a GenBank match.
    - genbank_unmatch_file: Contains GenBank records without a PubMed match.
    - pubmed_genbank_combined: A merged file of matched and unmatched data.
    """

    for idx, row in pubmed.iterrows():
        authors = row['Authors']
        if ',' in authors:
            authors = authors.split(',')
        else:
            authors = authors.split(';')
        first_author_surname = ''
        if authors:
            first_author = authors[0]
            first_author_name_list = first_author.split()
            if first_author_name_list[0].isupper():
                first_author_surname = ' '.join(first_author_name_list[1:])
            else:
                first_author_surname = ' '.join(first_author_name_list[:-1])
        pubmed.at[idx, 'FirstAuthorSurname'] = first_author_surname

        if str(row['PMID']).isdigit():
            short_name = f"{first_author_surname} ({row['Year']}, {row['PMID']})"
        else:
            short_name = f"{first_author_surname} ({row['Year']})"
        pubmed.at[idx, 'ShortName'] = short_name

    logger = virus_obj.get_logger('compare_matched')
    pubmed_match, genbank2pubmed, genbank_match, pubmed_unmatch, genbank_unmatch = match(
        virus_obj, pubmed, genbank_ref, logger)

    # won't use
    # summarize_complete_workflow_GPT(virus_obj, pubmed_match)
    # summarize_complete_workflow_GPT_or_R1(virus_obj, pubmed_match)

    pubmed_unmatch.to_excel(virus_obj.pubmed_unmatch_file)
    genbank_unmatch.to_excel(virus_obj.genbank_unmatch_file)

    logger = virus_obj.get_logger('compare_pubmed_only')
    logger.info('Pumbed only Literatures:', len(pubmed_unmatch))
    logger.report(summarize_pubmed_data(pubmed_unmatch))

    logger = virus_obj.get_logger('compare_genbank_only')
    logger.info('GenBank only References:', len(genbank_unmatch))
    logger.report(summarize_genbank_by_ref(genbank_unmatch))

    combined, columns = combine_file(
        pubmed_match, pubmed_unmatch, genbank_unmatch,
        genbank_feature, genbank_genes)

    combined.to_excel(str(virus_obj.pubmed_genbank_combined),
                      index=False, columns=columns)
    format_table(str(virus_obj.pubmed_genbank_combined))

    logger = virus_obj.get_logger('compare_matched')
    summarize_combined_data(
        virus_obj, combined, genbank_feature, genbank_genes, logger)

    logger.report(summarize_genbank_full_genome(
        genbank_match, genbank_feature,
        virus_obj.GENES))

    # yesno = input('Generate Chord diagram? [y/n]')
    # if yesno == 'y':
    #     from chord_diagram import gen_chord_diagram
    #     gen_chord_diagram(virus_obj, combined, genbank_feature)

    return pubmed, pubmed_match, genbank2pubmed


def match(virus, pubmed, genbank, logger):
    """
    Matches PubMed records with GenBank references based on PMID, accession numbers, and title similarity.
    Also processes hard-linked matches, logs match statistics, and returns matched and unmatched records.
    """

    match_by_pmid_list = []
    match_by_title_list = []
    match_by_acc_list = []

    matched_pub_id = []
    genbank_unmatch_list = {}

    pubmed['PMID'] = pubmed['PMID'].astype(str)
    genbank['PMID'] = genbank['PMID'].astype(str)

    for index, row in genbank.iterrows():
        pmid = row['PMID']
        pmid_list = [
            str(p).strip()
            for p in pmid.split(',')
        ]

        match_by_pmid = None

        # some times a PMID can match to multiple papers.
        if pmid:
            pubmed_paper = pubmed[pubmed['PMID'].isin(pmid_list)]

            if not pubmed_paper.empty:
                matched_pub_id.extend(pubmed_paper['PubID'].tolist())
                match_by_pmid = [row, pubmed_paper, row['RefID'], 'PMID']
                match_by_pmid_list.append(match_by_pmid)
            # else:
            #     genbank_unmatch_list[row['RefID']] = row
            # continue

        match_by_acc = None

        accession_list = row['accession']

        pubmed_paper = search_by_accession(pubmed, accession_list)
        if not pubmed_paper.empty:
            matched_pub_id.extend(pubmed_paper['PubID'].tolist())
            match_by_acc = [row, pubmed_paper, row['RefID'], 'ACCESSION']
            match_by_acc_list.append(match_by_acc)

        if match_by_acc or match_by_pmid:
            continue

        title = row['Title'].replace('Direct Submission', '')

        match_by_title = None

        if title:

            # Pubmed title always exists
            pubmed_paper = pubmed[pubmed['Title'].apply(
                lambda x:
                    (Levenshtein.distance(x.lower(), title.lower()) < 5) or
                    (title.lower() in x.lower()) or
                    (x.lower() in title.lower())
            )]

            if not pubmed_paper.empty:
                matched_pub_id.extend(pubmed_paper['PubID'].tolist())
                match_by_title = [row, pubmed_paper, row['RefID'], 'Title']
                match_by_title_list.append(match_by_title)

        if not match_by_title:
            genbank_unmatch_list[row['RefID']] = row

    # Given the hardlink excel, link the GenBank with literature
    # even when no match based on above criteria
    hard_link_list = []

    if virus.pubmed_genbank_hardlink:
        hard_link_df = pd.read_excel(str(virus.pubmed_genbank_hardlink))
        for idx, pair in hard_link_df.iterrows():
            pub_id = int(pair['PubID'])
            ref_id = int(pair['RefID'])
            pubmed_paper = pubmed[pubmed['PubID'].isin([pub_id])]
            genbank_ss = genbank[genbank['RefID'].isin([ref_id])]

            if genbank_ss.empty:
                continue

            genbank_ss = genbank_ss.iloc[0]

            hard_link_list.append([genbank_ss, pubmed_paper, ref_id, 'Hardlink'])
            matched_pub_id.extend(pubmed_paper['PubID'].tolist())

            if ref_id not in genbank_unmatch_list:
                print(f"Warning: hard link file is matching to a wrong RefID: {ref_id}")
            else:
                del genbank_unmatch_list[ref_id]

    genbank_match_list = (
        match_by_title_list +
        match_by_pmid_list +
        match_by_acc_list +
        hard_link_list
    )

    genbank_match_list = keep_most_related_match(genbank_match_list)

    dump_data = []
    for gen, pub_list, _, _ in genbank_match_list:
        gen = gen.copy()
        gen['Pub_PMID'] = ', '.join(pub_list['PMID'].tolist())
        dump_data.append(gen)

    pd.DataFrame(dump_data).to_excel(virus.output_excel_dir / f'{virus.name}_Genbank_match_PMID.xlsx')

    paired_pub_id_ref_id = []
    for _, paired_pubmed, ref_id, _ in genbank_match_list:
        for _, p_pubmed in paired_pubmed.iterrows():
            paired_pub_id_ref_id.append(
                (ref_id, p_pubmed['PubID'])
            )

    paired_pub_id_ref_id = [
        {
            'RefID': ref_id,
            'PubID': pub_id
        }
        for ref_id, pub_id in sorted(
            list(set(paired_pub_id_ref_id)))
    ]
    paired_pub_id_ref_id.sort(key=itemgetter('RefID', 'PubID'))
    pd.DataFrame(paired_pub_id_ref_id).to_csv(
        virus.paired_pub_id_ref_id_track, index=False)

    # TODO, should be a small data structure
    logger.info("Genbank match by pmid:",
                len(set(i[-2] for i in match_by_pmid_list)))
    logger.info("Genbank match by title:",
                len(set(i[-2] for i in match_by_title_list)))
    logger.info("Genbank match by acc:",
                len(set(i[-2] for i in match_by_acc_list)))
    logger.info("Genbank match by acc or title not by pubmed:",
                len(
                    set(i[-2] for i in (match_by_acc_list + match_by_title_list))
                    -
                    set(i[-2] for i in match_by_pmid_list)
                ))
    logger.info('Genbank match total:',
                len(set(i[-2] for i in genbank_match_list)))
    logger.info('-' * 80)

    pubmed_match = match_pubmed2genbank(genbank_match_list)

    genbank_match = {}
    for row, result, ref_id, method in genbank_match_list:
        genbank_match[ref_id] = row

    genbank_match = genbank_match.values()

    logger.info("Pubmed match by pmid:",
                len(match_pubmed2genbank(match_by_pmid_list)))
    logger.info("Pubmed match by title:",
                len(match_pubmed2genbank(match_by_title_list)))
    logger.info("Pubmed match by acc:",
                len(match_pubmed2genbank(match_by_acc_list)))
    logger.info('Pubmed match total:',
                len(pubmed_match))

    pubmed_unmatch = pubmed[~pubmed['PubID'].isin(matched_pub_id)]

    # Process Unmatched GenBank Records Using AI
    genbank_unmatch_list = pd.DataFrame(genbank_unmatch_list.values())

    return (
        pubmed_match,
        genbank_match_list,
        pd.DataFrame(genbank_match),
        pubmed_unmatch, genbank_unmatch_list)


def keep_most_related_match(genbank_match_list):
    genbank_link = defaultdict(list)
    for gen, publist, ref_id, method in genbank_match_list:
        for _, pub in publist.iterrows():
            genbank_link[ref_id].append((gen, pub, ref_id, method))

    keep_link = []
    discard_link = []
    method_order = ['PMID', 'Hardlink', 'ACCESSION', 'Title']
    for ref_id, links in genbank_link.items():
        for order in method_order:
            item = [
                i
                for i in links
                if i[-1] == order
            ]
            other = [
                i
                for i in links
                if i[-1] != order
            ]
            if item:
                keep_link.append(item[0])
                discard_link.extend(other + item[1:])
                break

    processed_pub_id = sorted([
        i[1]['PubID']
        for i in keep_link
    ])

    pubmed_link = defaultdict(list)
    for gen, pub, ref_id, method in discard_link:
        pub_id = pub['PubID']
        pubmed_link[pub_id].append((gen, pub, ref_id, method))

    # print(sorted(processed_pub_id))
    # print(pubmed_link.keys())
    for pub_id, links in pubmed_link.items():
        if pub_id in processed_pub_id:
            continue

        # print(pub_id, [i[-1] for i in links])
        for order in method_order:
            item = [
                i
                for i in links
                if i[-1] == order
            ]
            if item:
                keep_link.append(item[0])
                break

    ref_id_method_group = defaultdict(list)
    for (gen, pub, ref_id, method) in keep_link:
        ref_id_method_group[(ref_id, method)].append((gen, pub, ref_id, method))

    final_link = []
    for (ref_id, method), links in ref_id_method_group.items():
        pub_list = pd.DataFrame([
            i[1]
            for i in links
        ])

        final_link.append((links[0][0], pub_list, ref_id, method))

    return final_link


def search_by_accession(pubmed, accession_list):

    accession_list = [
        a.strip() for a in accession_list.split(',')
        if a.strip() and a.strip()[:2].upper() not in ['NC', 'NG', 'NM', 'NR']
    ]

    found = []

    for idx, row in pubmed.iterrows():
        accs = row['GenBank']
        accs = re.findall(r'[A-Z]{1,2}\d{5,7}', accs)
        if not accs:
            continue

        if len(set(accs) & set(accession_list)):
            found.append(row)
            # continue

        # acc1 = [
        #     i[:6]
        #     for i in accs
        # ]

        # acc2 = [
        #     i[:6]
        #     for i in accession_list
        # ]

        # counter1 = Counter(acc1)
        # counter2 = Counter(acc2)
        # overlap_count = sum(
        #     min(counter1[key], counter2[key])
        #     for key in counter1.keys() & counter2.keys())

        # if overlap_count > 1:
        #     found.append(row)

    return pd.DataFrame(found)


def match_pubmed2genbank(genbank_match):
    pubmed_match = defaultdict(dict)
    for g, pubmed_list, ref_id, method in genbank_match:
        for r, i in pubmed_list.iterrows():
            pubmed_match[r]['pubmed'] = i
            if 'genbank' not in pubmed_match[r]:
                pubmed_match[r]['genbank'] = []

            g = g.copy()
            g['match_method'] = method
            pubmed_match[r]['genbank'].append(g)

            # if 'method' not in pubmed_match[r]:
            #     pubmed_match[r]['method'] = []
            # pubmed_match[r]['method'].append(method)

    pubmed_match = [
        (v['pubmed'], v['genbank'])
        for v in pubmed_match.values()
    ]

    return pubmed_match


def summarize_combined_data(virus_obj, combined, features, genes, logger):
    summarize_report = []

    section = ['Summarize PubMed GenBank Match']
    summarize_report.append(section)

    genbank_only_pubmed = combined[
        (combined['match'] != 'Yes') &
        (combined['Reviewer(s) Seq'] == '') &
        (combined['PMID'] != '')
        ]
    section = ['GenBank only PMID, or not in first search PMID']
    section.append(len(genbank_only_pubmed))
    section.append(', '.join(genbank_only_pubmed['PMID'].to_list()))
    summarize_report.append(section)

    matches = combined[(combined['match'] == 'Yes')]

    section = ['Publish Year']
    publish_year = count_number([
        v for i, v in matches.iterrows()], 'Year', sorter=int_sorter)
    section.append(publish_year)

    publish_year = [
        int(v['Year']) for i, v in matches.iterrows()
        if v['Year'] and v['Year'] != 'NA']
    section.append(create_binnned_year(publish_year))
    summarize_report.append(section)

    # section = ['Journals']
    # journals = count_number([v for i, v in matches.iterrows()], 'Journal')
    # section.append(journals)
    # summarize_report.append(section)

    section = ['Seq method']
    methods = count_number(
        [v for i, v in matches.iterrows()], 'SeqMethod (PM)')
    section.append(methods)
    summarize_report.append(section)

    section = ['After matching, NumSeq from GenBank']
    accessions = set([
        j.strip()
        for i, v in matches.iterrows()
        for j in v['GenBank (GB)'].split(',')
        ])
    num_seq = len(accessions)
    section.append(num_seq)
    summarize_report.append(section)

    features = features[features['Accession'].isin(list(accessions))].copy()

    genes = genes[genes['Accession'].isin(list(accessions))]

    section = ['After match, GenBank empty sequence']
    genbank_column_map = {
        'Hosts': 'Host',
        'Specimen': 'isolate_source',
        'Countries': 'Country',
        'SampleYr': 'IsolateYear',
        'Genes': 'Genes',
    }
    for c in genbank_column_map.values():
        v = features[features[c] == '']
        section.append((c, 'num seq', len(v)))
    summarize_report.append(section)

    section = ['Pubmed Supplement GenBank']
    for name in [
            'Hosts', 'Specimen', 'SampleYr', 'Countries', 'Genes', 'SeqMethod']:

        count = 0
        count_acc = []
        for idx, row in matches.iterrows():
            p_value = row[f"{name} (PM)"]
            g_value = row[f"{name} (GB)"]
            g_value = re.sub(r"\s\(\d+\)", "", g_value)

            na_in_g_value = ('NA' in g_value) or (not g_value)

            p_value = [p for p in p_value.split(',') if p.strip() and p.strip().upper() != 'NA']
            g_value = [g for g in g_value.split(',') if g.strip() and g.strip().upper() != 'NA']

            if p_value and not g_value:
                count += 1

            if p_value and na_in_g_value:
                # accession may show in multiple rows, so calculate supplied accessions.
                accessions = set([
                    j.strip()
                    for j in row['GenBank (GB)'].split(',')
                    ])

                sub_features = features[features['Accession'].isin(list(accessions))]
                if name in genbank_column_map:
                    gb_column = genbank_column_map.get(name, name)
                    sub_features = sub_features[sub_features[gb_column] == '']
                    count_acc += sub_features['Accession'].tolist()
                else:
                    count_acc += accessions

        section.append((name, count))
        section.append((f"{name} num seq", len(set(count_acc))))
    summarize_report.append(section)

    # section = ['Similar virus']
    # section.append(summarize_similarity(combined, 'Viruses'))

    # section = ['Similar hosts']
    # section.append(summarize_similarity(combined, 'Hosts'))

    # section = ['Similar Specimens']
    # section.append(summarize_similarity(combined, 'Specimen'))

    # section = ['Similar countries']
    # section.append(summarize_similarity(combined, 'Countries'))

    # section = ['Similar Genes']
    # section.append(summarize_similarity(combined, 'Genes'))

    section = ['# Matched pubmed genbank']
    summarize_report.append(section)

    logger.report(summarize_report)

    logger.report(summarize_genbank_by_seq(virus_obj, features, genes, 'isolate_with_pubmed_'))


def summarize_similarity(df, col_name):

    count = 0
    for i, row in df.iterrows():
        pubmed = row[f'{col_name} (PM)']
        pubmed = [i.strip().lower() for i in pubmed.split(',') if i.strip().lower() != 'NA']

        genbank = row[f'{col_name} (GB)']
        genbank = get_values_of_value_count_list(genbank) if genbank else set()
        genbank = [i.lower() for i in genbank if i.lower() != 'NA']

        if (set(pubmed) & set(genbank)):
            count += 1

    return count


def combine_file(
        pubmed_match, pubmed_unmatch, genbank_unmatch,
        features_df, genes_df
        ):

    result = []
    for pubmed, genbank_list in pubmed_match:
        accessions = set([
             j.strip()
             for i in genbank_list
             for j in i['accession'].split(',')
        ])

        features = features_df[features_df['Accession'].isin(
            accessions)]

        genes = genes_df[genes_df['Accession'].isin(accessions)]

        features_stat = merge_feature_rows(features, genes)

        row = {
            'Authors': pubmed['Authors'],
            'Title': pubmed['Title'],
            'Journal': pubmed['Journal'],
            'Year': pubmed['Year'],
            'PMID': pubmed['PMID'],

            'Reviewer(s) Seq': pubmed['Reviewer(s) Seq'],
            'GPT seq (Y/N)': pubmed['GPT seq (Y/N)'],
            'Resolve Title': pubmed['Resolve Title'],
            'match': 'Yes',

            'Viruses (PM)': pubmed['Viruses'],
            'NumSeqs (PM)': pubmed['NumSeqs'],
            'Hosts (PM)': pubmed['Host'],
            'Specimen (PM)': pubmed['Specimen'],

            'SampleYr (PM)': pubmed['SampleYr'],
            'Countries (PM)': pubmed['Country'],

            'Genes (PM)': pubmed['Gene'],
            'SeqMethod (PM)': pubmed['SeqMethod'],
            'CloneMethod (PM)': pubmed['CloneMethod'],
            'GenBank (PM)': pubmed['GenBank'],

            'Viruses (GB)': features_stat['Organisms'],
            'NumSeqs (GB)': len(accessions),
            'Hosts (GB)': features_stat['Hosts'],
            'Specimen (GB)': features_stat['Specimens'],

            'SampleYr (GB)':  features_stat['IsolateYears'],
            'Countries (GB)': features_stat['Countries'],

            'Genes (GB)': features_stat['Gene'],
            'SeqMethod (GB)': '',
            'CloneMethod (GB)': '',
            'GenBank (GB)': ', '.join(sorted(list(accessions))),

            'NumSubSeqs': features_stat['NumSubSeqs'],
            'AlignLens (GB)': features_stat['AlignLens'],
            'PcntIDs (GB)': features_stat['PcntIDs'],
            # 'Combine Method': ','.join(list(method)),
        }

        result.append(row)

    for r, pubmed in pubmed_unmatch.iterrows():
        row = {
            'Authors': pubmed['Authors'],
            'Title': pubmed['Title'],
            'Journal': pubmed['Journal'],
            'Year': pubmed['Year'],
            'PMID': pubmed['PMID'],

            'Reviewer(s) Seq': pubmed['Reviewer(s) Seq'],
            'GPT seq (Y/N)': pubmed['GPT seq (Y/N)'],
            'Resolve Title': pubmed['Resolve Title'],

            'Viruses (PM)': pubmed['Viruses'],
            'NumSeqs (PM)': pubmed['NumSeqs'],
            'Hosts (PM)': pubmed['Host'],
            'Specimen (PM)': pubmed['Specimen'],
            'SampleYr (PM)': pubmed['SampleYr'],
            'Countries (PM)': pubmed['Country'],
            'Genes (PM)': pubmed['Gene'],
            'SeqMethod (PM)': pubmed['SeqMethod'],
            'CloneMethod (PM)': pubmed['CloneMethod'],
            'GenBank (PM)': pubmed['GenBank'],
            'Combine Method': ''
        }

        result.append(row)

    for row, genbank in genbank_unmatch.iterrows():

        accessions = set(genbank['accession'].split(','))
        features = features_df[features_df['Accession'].isin(
            accessions)]

        genes = genes_df[genes_df['Accession'].isin(accessions)]

        features_stat = merge_feature_rows(features, genes)

        row = {
            'Authors': genbank['Authors'],
            'Title': genbank['Title'],
            'Journal': genbank['Journal'],
            'Year': genbank['Year'],
            'PMID': genbank['PMID'],
            'Viruses (GB)': features_stat['Organisms'],
            'NumSeqs (GB)': len(accessions),
            'Hosts (GB)': features_stat['Hosts'],
            'Specimen (GB)': features_stat['Specimens'],
            'SampleYr (GB)': features_stat['IsolateYears'],
            'Countries (GB)': features_stat['Countries'],
            'Genes (GB)': features_stat['Gene'],
            'SeqMethod (GB)': '',
            'CloneMethod (GB)': '',
            'GenBank (GB)': ', '.join(sorted(list(accessions))),
            'NumSubSeqs': features_stat['NumSubSeqs'],
            'AlignLens (GB)': features_stat['AlignLens'],
            'PcntIDs (GB)': features_stat['PcntIDs'],
            'Combine Method': ''
        }

        result.append(row)

    columns = [
        'Authors',
        'Title',
        'Journal',
        'Year',
        'PMID',

        'Reviewer(s) Seq',
        'GPT seq (Y/N)',
        'Resolve Title',
        'match',

        'Viruses (PM)',
        'Viruses (GB)',

        'NumSeqs (PM)',
        'NumSeqs (GB)',

        'Hosts (PM)',
        'Hosts (GB)',

        'Specimen (PM)',
        'Specimen (GB)',

        'SampleYr (PM)',
        'SampleYr (GB)',

        'Countries (PM)',
        'Countries (GB)',

        'Genes (PM)',
        'Genes (GB)',

        'SeqMethod (PM)',
        'SeqMethod (GB)',

        'CloneMethod (PM)',
        'CloneMethod (GB)',

        'GenBank (PM)',
        'GenBank (GB)',

        'NumSubSeqs',
        'AlignLens (GB)',
        'PcntIDs (GB)',
        'Combine Method',
    ]

    for i in result:
        for c in columns:
            if c not in i:
                i[c] = ''

    return pd.DataFrame(result), columns


def format_table(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 20

    wb.save(excel_file)


def summarize_complete_workflow_GPT(virus_obj, pubmed_match):
    pubmed = pd.read_excel(virus_obj.pubmed_file, dtype=str).fillna('')

    likely = pubmed[
        (
            (pubmed['GPT (Y/N)'].str.lower().isin(('likely', 'unsure')))
        )
    ]
    logger = virus_obj.get_logger('pubmed_workflow')
    logger.info('GPT Title/Abstract likely:', len(likely))

    likely_with_seq = likely[(likely['GPT seq (Y/N)'].str.lower() == 'yes')]
    likely_wo_seq = likely[(likely['GPT seq (Y/N)'].str.lower() != 'yes')]
    logger.info('GPT Title/Abstract likely, GPT with sequence:', len(likely_with_seq))
    logger.info('GPT Title/Abstract likely, GPT wo sequence:', len(likely_wo_seq))

    two_reviwer_agree_GPT_with = likely_with_seq[
        (likely_with_seq['Resolve Seq'].str.lower() != 'no') &
        (
            (likely_with_seq['Reviewer(s) Seq'].str.lower() == 'yes') |
            (likely_with_seq['GPT seq (Y/N)'].str.lower() == 'yes')
        )
    ]
    logger.info(
        'GPT Title/Abstract likely, '
        'GPT with sequence, '
        'two reviewer agree with sequence',
        len(two_reviwer_agree_GPT_with))
    logger.info(
        'GPT Title/Abstract likely, '
        'GPT with sequence, '
        'no two reviewer agree with sequence',
        len(likely_with_seq) - len(two_reviwer_agree_GPT_with))

    two_reviwer_agree_GPT_wo = likely_wo_seq[
        (likely_wo_seq['Resolve Seq'].str.lower() != 'no') &
        (
            (likely_wo_seq['Reviewer(s) Seq'].str.lower() == 'yes') |
            (likely_wo_seq['GPT seq (Y/N)'].str.lower() == 'yes')
        )
    ]
    logger.info('GPT Title/Abstract likely, GPT wo sequence, two reviewer agree with sequence', len(two_reviwer_agree_GPT_wo))

    if virus_obj.pubmed_additional_from_gb:
        additional_pubmed = pd.read_excel(
            virus_obj.pubmed_additional_from_gb, dtype=str).fillna('')
    else:
        additional_pubmed = pd.DataFrame()
    logger.info('GPT Title/Abstract unlikely, GPT wo sequence, from GenBank only', len(additional_pubmed))

    if virus_obj.pubmed_search_missing:
        pubmed_missing = pd.read_excel(
            virus_obj.pubmed_search_missing, dtype=str).fillna('')
    else:
        pubmed_missing = pd.DataFrame()
    logger.info('Papers not found in PubMed and GenBank search:', len(pubmed_missing))

    unlikely = pubmed[
        (
            (pubmed['Reviewer1  (Y/N)'].str.lower() == 'unlikely')
            &
            (pubmed['GPT (Y/N)'].str.lower() == 'unlikely')
        )
    ]
    logger.info('In pubmed search, Title/Abstract unlikely, but in GenBank', len(unlikely))


def summarize_complete_workflow_GPT_or_R1(virus_obj, pubmed_match):
    pubmed = pd.read_excel(virus_obj.pubmed_file, dtype=str).fillna('')

    likely = pubmed[
        (
            (pubmed['Reviewer1  (Y/N)'].str.lower().isin(('likely', 'unsure')))
            |
            (pubmed['GPT (Y/N)'].str.lower().isin(('likely', 'unsure')))
        )
    ]

    logger = virus_obj.get_logger('pubmed_workflow')
    logger.info('GPT or R1 Title/Abstract likely:', len(likely))

    likely_with_seq = likely[(likely['GPT seq (Y/N)'].str.lower() == 'yes')]
    likely_wo_seq = likely[(likely['GPT seq (Y/N)'].str.lower() != 'yes')]
    logger.info('GPT or R1 Title/Abstract likely, GPT with sequence:', len(likely_with_seq))
    logger.info('GPT or R1 Title/Abstract likely, GPT wo sequence:', len(likely_wo_seq))

    two_reviwer_agree_GPT_with = likely_with_seq[
        (likely_with_seq['Resolve Seq'].str.lower() != 'no') &
        (
            (likely_with_seq['Reviewer(s) Seq'].str.lower() == 'yes') |
            (likely_with_seq['GPT seq (Y/N)'].str.lower() == 'yes')
        )
    ]
    logger.info(
        'GPT or R1 Title/Abstract likely, '
        'GPT with sequence, '
        'two reviewer agree with sequence',
        len(two_reviwer_agree_GPT_with))
    logger.info(
        'GPT or R1 Title/Abstract likely, '
        'GPT with sequence, '
        'no two reviewer agree with sequence',
        len(likely_with_seq) - len(two_reviwer_agree_GPT_with))

    two_reviwer_agree_GPT_wo = likely_wo_seq[
        (likely_wo_seq['Resolve Seq'].str.lower() != 'no') &
        (
            (likely_wo_seq['Reviewer(s) Seq'].str.lower() == 'yes') |
            (likely_wo_seq['GPT seq (Y/N)'].str.lower() == 'yes')
        )
    ]
    logger.info('GPT or R1 Title/Abstract likely, GPT wo sequence, two reviewer agree with sequence', len(two_reviwer_agree_GPT_wo))

    if virus_obj.pubmed_additional_from_gb:
        additional_pubmed = pd.read_excel(
            virus_obj.pubmed_additional_from_gb, dtype=str).fillna('')
    else:
        additional_pubmed = pd.DataFrame()
    logger.info('GPT or R1 Title/Abstract unlikely, GPT wo sequence, from GenBank only', len(additional_pubmed))

    if virus_obj.pubmed_search_missing:
        pubmed_missing = pd.read_excel(
            virus_obj.pubmed_search_missing, dtype=str).fillna('')
    else:
        pubmed_missing = pd.DataFrame()
    logger.info('Papers not found in PubMed and GenBank search:', len(pubmed_missing))

    unlikely = pubmed[
        (
            (pubmed['Reviewer1  (Y/N)'].str.lower() == 'unlikely')
            &
            (pubmed['GPT (Y/N)'].str.lower() == 'unlikely')
        )
    ]
    logger.info('In pubmed search, Title/Abstract unlikely, but in GenBank', len(unlikely))
