import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from DataFrameLogic import merge_feature_rows


def combine_file(
        pubmed_match, pubmed_unmatch, genbank_unmatch,
        features_df, genes_df
        ):

    result = []
    for pubmed, genbank_list in pubmed_match:
        accessions = set([
             j.strip()
             for i in genbank_list
             for j in i['accession'].split(',')
        ])

        features = features_df[features_df['Accession'].isin(
            accessions)]

        genes = genes_df[genes_df['Accession'].isin(accessions)]

        features_stat = merge_feature_rows(features, genes)

        row = {
            'Authors': pubmed['Authors'],
            'Title': pubmed['Title'],
            'Journal': pubmed['Journal'],
            'Year': pubmed['Year'],
            'PMID': pubmed['PMID'],

            'Reviewer(s) Seq': pubmed['Reviewer(s) Seq'],
            'GPT seq (Y/N)': pubmed['GPT seq (Y/N)'],
            'Resolve Title': pubmed['Resolve Title'],
            'match': 'Yes',

            'Viruses (PM)': pubmed['Viruses'],
            'NumSeqs (PM)': pubmed['NumSeqs'],
            'Hosts (PM)': pubmed['Host'],
            'Specimen (PM)': pubmed['IsolateType'],

            'SampleYr (PM)': pubmed['SampleYr'],
            'Countries (PM)': pubmed['Country'],
            'Genes (PM)': pubmed['Gene'],
            'SeqMethod (PM)': pubmed['SeqMethod'],
            'CloneMethod (PM)': pubmed['CloneMethod'],
            'GenBank (PM)': pubmed['GenBank'],

            'Viruses (GB)': features_stat['Organisms'],
            'NumSeqs (GB)': len(accessions),
            'Hosts (GB)': features_stat['Hosts'],

            'Specimen (GB)': features_stat['Specimens'],
            'SampleYr (GB)':  features_stat['IsolateYears'],
            'Countries (GB)': features_stat['Countries'],

            'Genes (GB)': features_stat['Gene'],
            'SeqMethod (GB)': '',
            'CloneMethod (GB)': '',
            'GenBank (GB)': ', '.join(sorted(list(accessions))),
            'AlignLens (GB)': features_stat['AlignLens'],
            'PcntIDs (GB)': features_stat['PcntIDs'],
        }

        result.append(row)

    for r, pubmed in pubmed_unmatch.iterrows():
        row = {
            'Authors': pubmed['Authors'],
            'Title': pubmed['Title'],
            'Journal': pubmed['Journal'],
            'Year': pubmed['Year'],
            'PMID': pubmed['PMID'],

            'Reviewer(s) Seq': pubmed['Reviewer(s) Seq'],
            'GPT seq (Y/N)': pubmed['GPT seq (Y/N)'],
            'Resolve Title': pubmed['Resolve Title'],

            'Viruses (PM)': pubmed['Viruses'],
            'NumSeqs (PM)': pubmed['NumSeqs'],
            'Hosts (PM)': pubmed['Host'],
            'Specimen (PM)': pubmed['IsolateType'],
            'SampleYr (PM)': pubmed['SampleYr'],
            'Countries (PM)': pubmed['Country'],
            'Genes (PM)': pubmed['Gene'],
            'SeqMethod (PM)': pubmed['SeqMethod'],
            'CloneMethod (PM)': pubmed['CloneMethod'],
            'GenBank (PM)': pubmed['GenBank'],
        }

        result.append(row)

    for row, genbank in genbank_unmatch.iterrows():

        accessions = set(genbank['accession'].split(','))
        features = features_df[features_df['Accession'].isin(
            accessions)]

        genes = genes_df[genes_df['Accession'].isin(accessions)]

        features_stat = merge_feature_rows(features, genes)

        row = {
            'Authors': genbank['Authors'],
            'Title': genbank['Title'],
            'Journal': genbank['Journal'],
            'Year': genbank['Year'],
            'PMID': genbank['PMID'],
            'Viruses (GB)': features_stat['Organisms'],
            'NumSeqs (GB)': len(accessions),
            'Hosts (GB)': features_stat['Hosts'],
            'Specimen (GB)': features_stat['Specimens'],
            'SampleYr (GB)': features_stat['IsolateYears'],
            'Countries (GB)': features_stat['Countries'],
            'Genes (GB)': features_stat['Gene'],
            'SeqMethod (GB)': '',
            'CloneMethod (GB)': '',
            'GenBank (GB)': ', '.join([
                i.strip().split('.')[0]
                for i in accessions
            ]),
            'AlignLens (GB)': features_stat['AlignLens'],
            'PcntIDs (GB)': features_stat['PcntIDs'],
        }

        result.append(row)

    columns = [
        'Authors',
        'Title',
        'Journal',
        'Year',
        'PMID',

        'Reviewer(s) Seq',
        'GPT seq (Y/N)',
        'Resolve Title',
        'match',

        'Viruses (PM)',
        'Viruses (GB)',

        'NumSeqs (PM)',
        'NumSeqs (GB)',

        'Hosts (PM)',
        'Hosts (GB)',

        'Specimen (PM)',
        'Specimen (GB)',

        'SampleYr (PM)',
        'SampleYr (GB)',

        'Countries (PM)',
        'Countries (GB)',

        'Genes (PM)',
        'Genes (GB)',

        'SeqMethod (PM)',
        'SeqMethod (GB)',

        'CloneMethod (PM)',
        'CloneMethod (GB)',

        'GenBank (PM)',
        'GenBank (GB)',

        'AlignLens (GB)',
        'PcntIDs (GB)',
    ]

    for i in result:
        for c in columns:
            if c not in i:
                i[c] = ''

    return pd.DataFrame(result), columns


def format_table(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 20

    wb.save(excel_file)
